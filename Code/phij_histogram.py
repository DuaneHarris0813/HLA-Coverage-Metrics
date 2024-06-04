#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Mar 11 14:02:17 2024

@author: duaneharris
"""


# IMPORT PACKAGES:
from math import ceil
from math import log10
from numpy import linspace
from openpyxl import load_workbook
import matplotlib.pyplot as plt


# DATA FILES:
fdr =  '../Results/'       # Results Folder
fs = ['HLA-A Phij.xlsx', # Results File, HLA-A
      'HLA-B Phij.xlsx', # Results File, HLA-B
      'HLA-C Phij.xlsx'] # Results File, HLA-C


# ALLELE TYPES:
at = ['HLA-A',
      'HLA-B',
      'HLA-C']


# PATHOGENS:
lb = ['Ebola GP1 (Zaire)',
      'Ebola GP1 (Sudan)',
      'SARS-CoV-2 Wuhan-Hu-1',
      'SARS-CoV-2 Delta AY.4',
      'SARS-CoV-2 Omicron BA.1',
      'SARS-CoV-2 Omicron BA.2',
      'SARS-CoV-2 Omicron BA.5']


# REGIONS:
rg = ['Australia', 
      'Europe',
      'North Africa',
      'North America',
      'North East Asia',
      'Oceania',
      'South and Central America',
      'South Asia',
      'South East Asia',
      'Sub-Saharan Africa',
      'Western Asia']


# IMMUNO-DOMINANT EPITOPES, EBOLA:
idpe = ['ATDVPSATK',
        'TDVPSATKR',
        'GFRSGVPPK',
        'AENCYNLEI',
        'RLASTVIYR',
        'TEDPSSGYY',
        'DTTIGEWAF',
        'TTIGEWAFW',
        'NQDGLICGL',
        'TELRTFSIL',
        'ALFCICKFV',
        'LFCICKFVF']


# IMMUNO-DOMINANT EPITOPES, SARS:
idps = ['YLQPRTFLL',
        'GVYFASTEK',
        'NLNESLIDL',
        'TLDSKTQSL',
        'RLQSLQTYV',
        'QIYKTPPIK']


# FIGURE SIZE OPTIONS:
fw  = 504   # Figure Width
fh  = 504/3 # Figure Height
ppi = 72    # Points Per Inch


# BIN OPTIONS:
nb = 50 # Number of Histogram Bins


#%% PARAMETERS:

    
# SIZES:
nt = len(fs) # Number of Allele Types
nd = len(lb) # Number of Pathogens
nr = len(rg) # Number of Regions  
    

# COLOR SCHEME (FOR IDE):
cr2 = [[0.0, 0.0, 1.0],
       [1.0, 0.5, 0.0],
       [0.0, 1.0, 0.0],
       [1.0, 0.0, 0.0],
       [1.0, 0.5, 0.5],
       [1.0, 0.0, 1.0],
       [1.0, 0.5, 1.0],
       [0.5, 0.5, 0.5],
       [1.0, 1.0, 0.0],
       [0.0, 1.0, 1.0],
       [0.0, 0.5, 1.0],
       [0.0, 0.0, 0.0]]


#%% READ DATA #################################################################


# INITIALIZE VARIABLES:
pn = [[[[] for j in range(nr)] for i in range(nd)] for t in range(nt)] # Peptide Name
sg = [[[[] for j in range(nr)] for i in range(nd)] for t in range(nt)] # Sigma Value


# Phi_j Results:
for t in range(nt):  
    r = load_workbook(fdr+fs[t])  
    for i in range(nd):   
        w = r[lb[i]] 
        for k in w.iter_rows(min_row=4,min_col=2,values_only=True):
            for j in range(nr):
                pn[t][i][j].append(k[0+2*j])
                sg[t][i][j].append(k[1+2*j])
                
                            
#%% DETERMINE MAX VALUE #######################################################


mx = 0
for t in range(nt):
    for i in range(nd):
        for j in range(nr):
            mx = max(max(sg[t][i][j]),mx)
Mx = ceil(mx*100)/100
bn = linspace(0,Mx,nb+1)
              
                
#%% HISTOGRAMS ################################################################


for t in range(nt):
    for i in range(nd):
        for j in range(nr):
                 
            if i<=1:
                idp = idpe
            else:
                idp = idps
             
            fig,ax1 = plt.subplots(figsize=(fw/ppi,fh/ppi),dpi=300)
            ax2 = ax1.twinx()
            (n,bins,patches)=ax1.hist(sg[t][i][j],bins=bn,edgecolor='black',weights=[1/len(sg[t][i][j])]*len(sg[t][i][j]))
            
            for k in range(len(idp)):
                ix = [l for l in range(len(pn[t][i][j])) if pn[t][i][j][l]==idp[k]]
                if len(ix)>0:
                    ix = ix[0]
                    
                    ind = list(bins >= sg[t][i][j][ix]).index(True) - 1
                    
                    s0 = str(n[ind])
                    s1 = s0[2:]
                    s2 = s1.lstrip('0')
                    ex = len(s1)-len(s2)+1
                    np = log10(n[ind]*10**(ex))-ex

                    h1 = np+0.1
                    h2 = h1+0.25
                    
                    ax2.plot([sg[t][i][j][ix],sg[t][i][j][ix]],[h1,h2],label=idp[k],color=cr2[k])

            ax1.set_yscale('log')
            ax1.set_ylim((1e-4,1e1))
            ax1.set_yticks([1e-4,1e-3,1e-2,1e-1,1e0,1e1])
            ax2.set_ylim((-4,1))
            ax2.set_yticks([])
            
            plt.title(at[t])
                     
            if t==2:
                ax1.set_xlabel(r'$\phi_j$')
            
            if t!=2:
                ax1.tick_params(labelbottom=False)
            
            if t==2:
                plt.legend(fontsize=8, bbox_to_anchor=(0,-0.25), loc='upper left', ncol=4)
                                
            plt.savefig(fdr+'Figures/Phi Histograms/'+lb[i]+'/'+rg[j]+'/'+at[t]+' Phi Histogram.pdf',format='pdf',bbox_inches='tight')
            plt.show()   