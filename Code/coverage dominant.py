#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Feb 26 11:00:19 2024

@author: duaneharris
"""

# IMPORT PACKAGES:
from openpyxl import load_workbook
import hla_functions as hf


# DATA FOLDER:
fd  = '../Data/'
fdr = '../Results/'
fr  = 'Coverage Results.xlsx'


# DATA FILES:
fa =  'HLA-A Allele Frequencies.xlsx'               # File for Allele Frequency
fs = ['HLA-A_Ebola_Zaire_GP1_Weighted_Results.xlsx',
      'HLA-A_Ebola_Sudan_GP1_Weighted_Results.xlsx',      
      'HLA-A_SARS_Wuhan-Hu-1_Weighted_Results.xlsx',     
      'HLA-A_SARS_DeltaAY4_Weighted_Results.xlsx',
      'HLA-A_SARS_OmicronBA1_Weighted_Results.xlsx',
      'HLA-A_SARS_OmicronBA2_Weighted_Results.xlsx',
      'HLA-A_SARS_OmicronBA5_Weighted_Results.xlsx']
lb = ['Ebola GP1 (Zaire)',
      'Ebola GP1 (Sudan)',
      'SARS-CoV-2 Wuhan-Hu-1',
      'SARS-CoV-2 Delta AY.4',
      'SARS-CoV-2 Omicron BA.1',
      'SARS-CoV-2 Omicron BA.2',
      'SARS-CoV-2 Omicron BA.5']


# fa =  'HLA-B Allele Frequencies.xlsx'               # File for Allele Frequency
# fs = ['HLA-B_Ebola_Zaire_GP1_Weighted_Results.xlsx',
#       'HLA-B_Ebola_Sudan_GP1_Weighted_Results.xlsx',      
#       'HLA-B_SARS_Wuhan-Hu-1_Weighted_Results.xlsx',      
#       'HLA-B_SARS_DeltaAY4_Weighted_Results.xlsx',
#       'HLA-B_SARS_OmicronBA1_Weighted_Results.xlsx',
#       'HLA-B_SARS_OmicronBA2_Weighted_Results.xlsx',
#       'HLA-B_SARS_OmicronBA5_Weighted_Results.xlsx']
# lb = ['Ebola GP1 (Zaire)',
#       'Ebola GP1 (Sudan)',
#       'SARS-CoV-2 Wuhan-Hu-1',
#       'SARS-CoV-2 Delta AY.4',
#       'SARS-CoV-2 Omicron BA.1',
#       'SARS-CoV-2 Omicron BA.2',
#       'SARS-CoV-2 Omicron BA.5']


# fa =  'HLA-C Allele Frequencies.xlsx'               # File for Allele Frequency
# fs = ['HLA-C_Ebola_Zaire_GP1_Weighted_Results.xlsx',
#       'HLA-C_Ebola_Sudan_GP1_Weighted_Results.xlsx',      
#       'HLA-C_SARS_Wuhan-Hu-1_Weighted_Results.xlsx',      
#       'HLA-C_SARS_DeltaAY4_Weighted_Results.xlsx',
#       'HLA-C_SARS_OmicronBA1_Weighted_Results.xlsx',
#       'HLA-C_SARS_OmicronBA2_Weighted_Results.xlsx',
#       'HLA-C_SARS_OmicronBA5_Weighted_Results.xlsx']
# lb = ['Ebola GP1 (Zaire)',
#       'Ebola GP1 (Sudan)',
#       'SARS-CoV-2 Wuhan-Hu-1',
#       'SARS-CoV-2 Delta AY.4',
#       'SARS-CoV-2 Omicron BA.1',
#       'SARS-CoV-2 Omicron BA.2',
#       'SARS-CoV-2 Omicron BA.5']


#%% PARAMETERS ################################################################
    

# REGIONS:
rg = ['Australia', 
      'Europe',
      'North Africa',
      'North America',
      'North East Asia',
      'Oceania',
      'South and Central America',
      'South Asia',
      'South East Asia',
      'Sub-Saharan Africa',
      'Western Asia']


idpe = ['ATDVPSATK',
        'TDVPSATKR',
        'GFRSGVPPK',
        'AENCYNLEI',
        'RLASTVIYR',
        'TEDPSSGYY',
        'DTTIGEWAF',
        'TTIGEWAFW',
        'NQDGLICGL',
        'TELRTFSIL',
        'ALFCICKFV',
        'LFCICKFVF']
idps = ['YLQPRTFLL',
        'GVYFASTEK',
        'NLNESLIDL',
        'TLDSKTQSL',
        'RLQSLQTYV',
        'QIYKTPPIK']


# GET POSITION WEIGHTS AND ENRICHMENT SCORES:
pw = hf.compute_position_weights()
es = hf.compute_enrichment_scores()


# GET ALLELE FREQUENCY DATA:
a  = load_workbook(fd+fa)
rs = a.sheetnames
nr = len(rs)
af = [{} for j in range(nr)]
for j in range(nr):
    af[j],S = hf.allele_frequency_dict(a[rs[j]])
        
        
#%% COMPUTE SCORES ############################################################


# DETERMINE NUMBER OF DISEASES:
if isinstance(fs,str): # If fs is a single string
    nd = 1             # Then the number of diseases is 1
else:                  # If fs is a list
    nd = len(fs)       # Then the number of diseases is length of list

       
# INITIALIZE COVERAGE VARIABLE:      
F = [[] for i in range(nd)]
 

# COMPUTE COVERAGE RATIOS:
for i in range(nd): # Iterate Through Diseases 


    # LOAD DATA FOR DISEASE i:
    if isinstance(fs,str): d = load_workbook(fd+fs)   
    else: d = load_workbook(fd+fs[i])
    
    
    # READ NECESSARY DATA:
    if i<=1: idp=idpe
    else: idp=idps
    
        
    for j in range(nr): # Iterate Through Regions
    
    
        # GET DATA FOR REGION j:
        a1 = []; p1 = []; b1 = [];
        for k in d[rg[j]].iter_rows(min_row=2, min_col=1, max_col=7, values_only=True):
            a1.append(k[0])     # Allele for Disease i Region j
            p1.append(k[5])     # Peptides for Disease i Region j
            b1.append(k[6])     # Binding Affinities for Disease i Region j
   
   
        # COMPUTE F RATIO:     
        F[i].append(hf.compute_F(a1,p1,b1,af[j],es,pw,idp))
        
        
#%% SAVE RESULTS:


# LOAD RESULTS SHEET:
r = load_workbook(fdr+fr)
w = r['Dominant Coverage']


# DETERMINE ALLELE TYPE:
at = fa[4]
if at=='A':
    rrw = 4
elif at=='B':
    rrw = 18
elif at=='C':
    rrw = 32
    
    
# WRITE TO FILE:
for i in range(nd):
    for j in range(nr):
        w.cell(rrw+j,i+3).value = F[i][j]


# SAVE FILE:
r.save(fdr+fr)