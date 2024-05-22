#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Oct 10 16:18:13 2023

@author: duaneharris
"""


# IMPORT PACKAGES:
from Bio import SeqIO
from numpy import unique
from openpyxl import load_workbook
import hla_functions as hf


# DATA FOLDERS AND FILES:
fd  = '../Data/'              # Data Folder
fdr = '../Results/'           # Results Folder
fp  = 'Sequences.fasta'       # Protein Sequence File
fr1 = 'Coverage Results.xlsx' # Regional Results File
fr2 = 'Individual Results'    # Individual Results File


# HLA SPECIFIC FILES AND LABLES:

fa =  'HLA-A Allele Frequencies.xlsx'                  # Allele Frequency File
fs = ['HLA-A_Ebola_Zaire_GP1_Weighted_Results.xlsx',   # Pathogen Files
      'HLA-A_Ebola_Sudan_GP1_Weighted_Results.xlsx',      
      'HLA-A_Ebola_Zaire_NP_Weighted_Results.xlsx',
      'HLA-A_Ebola_Sudan_NP_Weighted_Results.xlsx',
      'HLA-A_SARS_Wuhan-Hu-1_Weighted_Results.xlsx',      
      'HLA-A_SARS_DeltaAY4_Weighted_Results.xlsx',
      'HLA-A_SARS_OmicronBA1_Weighted_Results.xlsx',
      'HLA-A_SARS_OmicronBA2_Weighted_Results.xlsx',
      'HLA-A_SARS_OmicronBA5_Weighted_Results.xlsx',
      'HLA-A_Burkholderia_HCP1_Weighted_Results.xlsx']
lb = ['Ebola GP1 (Zaire)',                             # Pathogen Sheet Names
      'Ebola GP1 (Sudan)',
      'Ebola NP (Zaire)',
      'Ebola NP (Sudan)',
      'SARS-CoV-2 Wuhan-Hu-1',
      'SARS-CoV-2 Delta AY.4',
      'SARS-CoV-2 Omicron BA.1',
      'SARS-CoV-2 Omicron BA.2',
      'SARS-CoV-2 Omicron BA.5', 
      'Burkholderia HCP1']

# fa =  'HLA-B Allele Frequencies.xlsx'                  # Allele Frequency File
# fs = ['HLA-B_Ebola_Zaire_GP1_Weighted_Results.xlsx',   # Pathogen Files
#       'HLA-B_Ebola_Sudan_GP1_Weighted_Results.xlsx',      
#       'HLA-B_Ebola_Zaire_NP_Weighted_Results.xlsx',
#       'HLA-B_Ebola_Sudan_NP_Weighted_Results.xlsx',
#       'HLA-B_SARS_Wuhan-Hu-1_Weighted_Results.xlsx',      
#       'HLA-B_SARS_DeltaAY4_Weighted_Results.xlsx',
#       'HLA-B_SARS_OmicronBA1_Weighted_Results.xlsx',
#       'HLA-B_SARS_OmicronBA2_Weighted_Results.xlsx',
#       'HLA-B_SARS_OmicronBA5_Weighted_Results.xlsx',
#       'HLA-B_Burkholderia_HCP1_Weighted_Results.xlsx']
# lb = ['Ebola GP1 (Zaire)',                             # Pathogen Sheet Names
#       'Ebola GP1 (Sudan)',
#       'Ebola NP (Zaire)',
#       'Ebola NP (Sudan)',
#       'SARS-CoV-2 Wuhan-Hu-1',
#       'SARS-CoV-2 Delta AY.4',
#       'SARS-CoV-2 Omicron BA.1',
#       'SARS-CoV-2 Omicron BA.2',
#       'SARS-CoV-2 Omicron BA.5', 
#       'Burkholderia HCP1']

# fa =  'HLA-C Allele Frequencies.xlsx'                  # Allele Frequency File
# fs = ['HLA-C_Ebola_Zaire_GP1_Weighted_Results.xlsx',   # Pathogen Files
#       'HLA-C_Ebola_Sudan_GP1_Weighted_Results.xlsx',      
#       'HLA-C_Ebola_Zaire_NP_Weighted_Results.xlsx',
#       'HLA-C_Ebola_Sudan_NP_Weighted_Results.xlsx',
#       'HLA-C_SARS_Wuhan-Hu-1_Weighted_Results.xlsx',      
#       'HLA-C_SARS_DeltaAY4_Weighted_Results.xlsx',
#       'HLA-C_SARS_OmicronBA1_Weighted_Results.xlsx',
#       'HLA-C_SARS_OmicronBA2_Weighted_Results.xlsx',
#       'HLA-C_SARS_OmicronBA5_Weighted_Results.xlsx',
#       'HLA-C_Burkholderia_HCP1_Weighted_Results.xlsx']
# lb = ['Ebola GP1 (Zaire)',                             # Pathogen Sheet Names
#       'Ebola GP1 (Sudan)',
#       'Ebola NP (Zaire)',
#       'Ebola NP (Sudan)',
#       'SARS-CoV-2 Wuhan-Hu-1',
#       'SARS-CoV-2 Delta AY.4',
#       'SARS-CoV-2 Omicron BA.1',
#       'SARS-CoV-2 Omicron BA.2',
#       'SARS-CoV-2 Omicron BA.5', 
#       'Burkholderia HCP1']


# OPTIONS:
ta  = 25 # Top Alleles (Max 25)


#%% PARAMETERS ################################################################


# GET POSITION WEIGHTS AND ENRICHMENT SCORES:
pw = hf.compute_position_weights()  # Position Weights
es = hf.compute_enrichment_scores() # Enrichment Scores


# GET ALLELE FREQUENCY DATA:
a  = load_workbook(fd+fa)    # Load Allele Frequecy File 
rs = a.sheetnames            # Region Names
nr = len(rs)                 # Number of Regions
af = [{} for j in range(nr)] # Initialize Allele Frequency Dictionary
Z  = []                      # Initialize Allele Frequency Sums
for j in range(nr):                              # Iterate Through Regions
    af[j],S = hf.allele_frequency_dict(a[rs[j]]) # Get Frequencies for Region j
    Z.append(S)                                  # Sum of Frequencies for Region j
        
        
#%% LOAD PROTEIN SEQUENCE DATA ################################################


# CREATE PEPTIDE SEQUENCE DICTIONARY:
pd = {}                                     # Initialize Protein Dictionary
for i in SeqIO.parse(fd+fp,format='fasta'): # Iterate Through FASTA Entries
    pd[i.description] = str(i.seq)          # Add Sequence i to Dictionary


# DETERMINE NUMBER OF DISEASES:
if isinstance(fs,str): nd = 1  # If fs is a string, then nd is 1
else: nd = len(fs)             # If fs is a list, then nd is its length
    

# RETRIEVE FULL PROTEIN SEQUENCES:
ep = [[] for i in range(nd)]           # Initialize Epitope Variable
for i in range(nd):                    # Iterate Through Pathogens
    for j in range(len(pd[lb[i]])-8):  # Iterate Through Regions
        ep[i].append(pd[lb[i]][j:j+9]) # Add Epitope j for Pathogen i to List


#%% COMPUTE SCORES ############################################################

       
# INITIALIZE VARIABLES:      
an = [[] for j in range(nr)]                      # Allele Names
f1 = [[] for j in range(nr)]                      # Allele Frequencies
f2 = []                                           # Individual Frequencies
sg = [[[] for j in range(nr)] for i in range(nd)] # Sigma Values
c1 = [[] for i in range(nd)]                      # Regional Coverage Scores
c2 = [[] for i in range(nd)]                      # Individual Coverage Scores


for i in range(nd): # Iterate Through Diseases 


    # LOAD DATA FOR DISEASE i:
    if isinstance(fs,str): d = load_workbook(fd+fs)   
    else: d = load_workbook(fd+fs[i])
    
        
    for j in range(nr): # Iterate Through Regions
    
    
        # GET DATA FOR REGION j:
        a1 = []; b1 = []; p1 = [] # Initialize Variables
        for k in d[rs[j]].iter_rows(min_row=2, values_only=True):
            if isinstance(k[0],str): # If the entry is not blank, Then
                a1.append(k[0]) # Allele Name
                b1.append(k[6]) # Binding Affinity
                p1.append(k[5]) # Peptide Name
        MN = len(a1) # Total Number of Alleles/Peptides
        
        
        # CHECK DATA FOR ERRORS:
        a2 = list(af[j].keys())[0:ta] # Allele Names for Top ta Alleles
        au = list(unique(a1))
        for k in range(len(au)):
            if au[k] not in a2:
                print('Allele Mismatch '+lb[i]+', '+rs[j]+', '+au[k])
        for k in range(len(a2)):
            if a2[k] not in au:
                print('Allele Mismatch '+lb[i]+', '+rs[j]+', '+a2[k])  
        pu = list(unique(p1))
        for k in range(len(pu)):
            if pu[k] not in ep[i]:
                print('Epitope Mismatch, '+lb[i]+', '+rs[j]+', '+pu[k])
        for k in range(len(ep[i])):
            if ep[i][k] not in pu:
                print('Epitope Mismatch, '+lb[i]+', '+rs[j]+', '+ep[i][k])
                
                
        # HLA-C Australia has only 22 alleles. If ta>22, then we will need to 
        # append blank entries onto a2 so that the rest of the code will run.
        if fa[4]=='C' and j==0 and len(a2)<ta: 
            for k in range(len(a2),ta):
                a2.append('')
        an[j] = a2 # Allele Names for Region j
        
        
        # COMPUTE ALLELE AND INDIVIDUAL FREQUENCIES:
        if i==0:
            for k in range(ta):
                if a2[k]=='':
                    f1[j].append(-1)
                else:
                    f1[j].append(af[j][a2[k]])
            f2.append(hf.compute_rho(f1[j]))
            for k in range(ta):
                if f1[j][k]==-1: f1[j][k]=0
                
                
        # COMPUTE SIGMA FOR EACH ALLELE:
        for k in range(ta):
            if a2[k] == '':
                sg[i][j].append(0)
            else:
                ia = [l for l in range(MN) if a1[l] == a2[k]]
                sg[i][j].append(hf.compute_sigma([b1[l] for l in ia],[p1[l] for l in ia],es,pw))
                
                
        # COMPUTE REGIONAL COVERAGE SCORE:
        c1[i].append(hf.compute_C(a1,b1,p1,af[j],es,pw))
        
        
        # COMPUTE INDIVIDUAL COVERAGE SCORES:
        c2[i].append(hf.compute_I(sg[i][j]))


#%% WRITE RESULTS TO FILE #####################################################


# LOAD RESULTS SHEETS:
r = load_workbook(fdr+fr1)


# ALLELE TYPE:
at = fa[4]


# COVERAGE RESULTS:
w = r['Coverage']
if at == 'A':
    rrw = 4
elif at == 'B':
    rrw = 18
elif at == 'C':
    rrw = 32
for i in range(nd):
    for j in range(nr):
        w.cell(rrw+j,i+3).value = c1[i][j]
        
        
# FREQ VS SIGMA RESULTS:
if at == 'A':
    rrw = 4
elif at == 'B':
    rrw = 33
elif at == 'C':
    rrw = 62
for j in range(nr):
    w = r[rs[j]]
    for i in range(nd):
        for k in range(ta):
            if i==0:
                w.cell(rrw+k,2).value = an[j][k]
                w.cell(rrw+k,3).value = f1[j][k]
            w.cell(rrw+k,4+i).value = sg[i][j][k]
        
        
# SAVE RESULTS 1:
r.save(fdr+fr1)


# INDIVIDUAL RESULTS:
for j in range(nr):   
    
    # LOAD RESULT FILE:
    r = load_workbook(fdr+fr2+' '+rs[j]+'.xlsx')
     
    # FREQUENCY:
    w = r['Frequency']
    for k in range(ta):
        w.cell(rrw+k,2).value = an[j][k]
        w.cell(rrw-1,3+k).value = an[j][k]
    for k in range(ta):
        for l in range(ta):
            w.cell(rrw+k,3+l).value = f2[j][k][l]
            
    # COVERAGE:
    for i in range(nd):
        w = r[lb[i]]
        for k in range(ta):
            w.cell(rrw+k,2).value = an[j][k]
            w.cell(rrw-1,3+k).value = an[j][k]
        for k in range(ta):
            for l in range(ta):
                w.cell(rrw+k,3+l).value = c2[i][j][k][l]
    
    # SAVE RESULTS 2
    r.save(fdr+fr2+' '+rs[j]+'.xlsx')        
