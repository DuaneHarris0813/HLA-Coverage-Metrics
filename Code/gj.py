#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Mar 12 11:08:10 2024

@author: duaneharris
"""


# IMPORT PACKAGES:
from Bio import SeqIO
from openpyxl import load_workbook
import hla_functions as hf


# DATA FILES:
fd  = '../Data/'        # Data Folder
fdr = '../Results/'     # Results Folder
fp  = 'Sequences.fasta' # Protein Sequence File
fr  = 'Gj.xlsx'         # Results File


#%% PARAMETERS ################################################################
    

# PATHOGEN LABELS / SHEET NAMES:
lb = ['Ebola GP1 (Zaire)',                             # Pathogen Sheet Names
      'Ebola GP1 (Sudan)',
      'SARS-CoV-2 Wuhan-Hu-1',
      'SARS-CoV-2 Delta AY.4',
      'SARS-CoV-2 Omicron BA.1',
      'SARS-CoV-2 Omicron BA.2',
      'SARS-CoV-2 Omicron BA.5', 
      'Burkholderia HCP1']


# GET POSITION WEIGHTS AND ENRICHMENT SCORES:
pw = hf.compute_position_weights()  # Position Weights
es = hf.compute_enrichment_scores() # Enrichment Scores


# SIZES:
nd = len(lb) # Number of Pathogens


#%% LOAD PROTEIN SEQUENCE DATA ################################################


# CREATE PEPTIDE SEQUENCE DICTIONARY:
pd = {}                                     # Initialize Protein Dictionary
for i in SeqIO.parse(fd+fp,format='fasta'): # Iterate Through FASTA Entries
    pd[i.description] = str(i.seq)          # Add Sequence i to Dictionary
    

# RETRIEVE FULL PROTEIN SEQUENCES:
ep = [[] for i in range(nd)]           # Initialize Epitope Variable
for i in range(nd):                    # Iterate Through Pathogens
    for j in range(len(pd[lb[i]])-8):  # Iterate Through Regions
        ep[i].append(pd[lb[i]][j:j+9]) # Add Epitope j for Pathogen i to List
        
    
#%% MAIN ######################################################################    


gj  = []
gjo = [[] for i in range(nd)]
epo = [[] for i in range(nd)]
for i in range(nd):   
    gj.append(hf.compute_g(ep[i],es,pw))
    gjo[i],epo[i] = zip(*sorted(zip(gj[i],ep[i]),reverse=True))


#%% WRITE DATA TO FILE ########################################################


r = load_workbook(fdr+fr)
for i in range(nd):
    for j in range(len(gjo[i])):  
        r['Gj'].cell(4+j,2+2*i).value = epo[i][j]
        r['Gj'].cell(4+j,3+2*i).value = gjo[i][j]       
r.save(fdr+fr)