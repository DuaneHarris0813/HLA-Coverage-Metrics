#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Mar 12 11:47:25 2024

@author: duaneharris
"""


# IMPORT PACKAGES:
from math import ceil
from math import floor
from numpy import linspace
from openpyxl import load_workbook
import matplotlib.pyplot as plt


# DATA FILES:
fdr =  '../Results/' # Results Folder
fs  = 'Gj.xlsx'      # Results File


# PATHOGENS:
lb = ['Zaire Ebola GP',
      'Sudan Ebola GP',
      'SARS-CoV-2 Wuhan-Hu-1',
      'SARS-CoV-2 Delta AY.4',
      'SARS-CoV-2 Omicron BA.1',
      'SARS-CoV-2 Omicron BA.2',
      'SARS-CoV-2 Omicron BA.5']


# IMMUNO-DOMINANT PEPTIDES, EBOLA:
idpe = ['ATDVPSATK',
        'TDVPSATKR',
        'GFRSGVPPK',
        'AENCYNLEI',
        'RLASTVIYR',
        'TEDPSSGYY',
        'DTTIGEWAF',
        'TTIGEWAFW',
        'NQDGLICGL',
        'TELRTFSIL',
        'ALFCICKFV',
        'LFCICKFVF']


# IMMUNO-DOMINANT PEPTIDES, SARS:
idps = ['YLQPRTFLL',
        'GVYFASTEK',
        'NLNESLIDL',
        'TLDSKTQSL',
        'RLQSLQTYV',
        'QIYKTPPIK']


# FIGURE SIZE:
fw  = 504   # Figure Width
fh  = 504/3 # Figure Height
ppi = 72    # Points Per Inch (Conversion Ratio)


# SIZE OPTIONS:
nd = 7  # Number of Diseases
nb = 50 # Number of Bins


# COLOR SCHEME (FOR IDE):
cr2 = [[0.0, 0.0, 1.0],
       [1.0, 0.5, 0.0],
       [0.0, 1.0, 0.0],
       [1.0, 0.0, 0.0],
       [1.0, 0.5, 0.5],
       [1.0, 0.0, 1.0],
       [1.0, 0.5, 1.0],
       [0.5, 0.5, 0.5],
       [1.0, 1.0, 0.0],
       [0.0, 1.0, 1.0],
       [0.0, 0.5, 1.0],
       [0.0, 0.0, 0.0]]


#%% READ DATA #################################################################


# INITIALIZE VARIABLES:
pn = [[] for i in range(nd)]
gj = [[] for i in range(nd)]


# IMMUNOGENICITY RESULTS:
r = load_workbook(fdr+fs)
for i in r['Gj'].iter_rows(min_row=4,min_col=2,max_col=15,values_only=True):
    for j in range(nd):
        if isinstance(i[0+2*j],str):
            pn[j].append(i[0+2*j])
            gj[j].append(i[1+2*j])
   
            
#%% DETERMINE MAX VALUE #######################################################


mn = 1; mx = 0
for i in range(nd):
    mn = min(min(gj[i]),mn)
    mx = max(max(gj[i]),mx)
Mn = floor(mn*100)/100
Mx = ceil(mx*100)/100
bn = linspace(Mn,Mx,nb+1)


#%% CHANGE LABEL NAMES ########################################################


# PAHTOGEN NAMES: 
lbl = lb[:]
for i in range(nd):
    lbl[i] = lbl[i].replace('GP1','GP')


#%% HISTOGRAMS ################################################################


for i in range(nd):
    
    
    if i<=1:
        idp = idpe
    else:
        idp = idps
        
    
    fig,ax = plt.subplots(figsize=(fw/ppi,fh/ppi),dpi=300)
    (n,bins,patches) = ax.hist(gj[i],bins=bn,edgecolor='black',weights=[1/len(gj[i])]*len(gj[i]))
    plt.ylim((0,0.1))
    
    for k in range(len(idp)):
        ix = [l for l in range(len(pn[i])) if pn[i][l]==idp[k]]
        if len(ix)>0:
            ix = ix[0]
            
            ind = list(bins >= gj[i][ix]).index(True) - 1
            h1 = n[ind]+0.002
            h2 = h1+0.005
            
            plt.plot([gj[i][ix],gj[i][ix]],[h1,h2],label=idp[k],color=cr2[k])
            
    plt.xlabel(r'$g_j$')
    plt.title(lbl[i])
    plt.legend(fontsize=8, bbox_to_anchor=(0,-0.25), loc='upper left', ncol=4)
              
    plt.savefig(fdr+'Figures/G Histograms/'+lb[i]+' G Histogram.pdf',format='pdf',bbox_inches='tight')
    plt.show()