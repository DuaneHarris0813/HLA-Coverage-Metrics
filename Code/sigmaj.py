#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Mar  8 11:33:11 2024

@author: duaneharris
"""


# IMPORT PACKAGES:
from numpy import exp
from numpy import unique
from openpyxl import load_workbook
import hla_functions as hf


# DATA FILES:
fd  = '../Data/'
fdr = '../Results/'


# OPTIONAL DATA FILES:


# fa =  'HLA-A Allele Frequencies.xlsx'               # File for Allele Frequency
# fs = ['HLA-A_Ebola_Zaire_GP1_Weighted_Results.xlsx',
#       'HLA-A_Ebola_Sudan_GP1_Weighted_Results.xlsx',      
#       'HLA-A_Ebola_Zaire_NP_Weighted_Results.xlsx',
#       'HLA-A_Ebola_Sudan_NP_Weighted_Results.xlsx',
#       'HLA-A_SARS_Wuhan-Hu-1_Weighted_Results.xlsx',      
#       'HLA-A_SARS_DeltaAY4_Weighted_Results.xlsx',
#       'HLA-A_SARS_OmicronBA1_Weighted_Results.xlsx',
#       'HLA-A_SARS_OmicronBA2_Weighted_Results.xlsx',
#       'HLA-A_SARS_OmicronBA5_Weighted_Results.xlsx',
#       'HLA-A_Burkholderia_HCP1_Weighted_Results.xlsx']
# fr =  'HLA-A Sigmaj.xlsx'


# fa =  'HLA-B Allele Frequencies.xlsx'               # File for Allele Frequency
# fs = ['HLA-B_Ebola_Zaire_GP1_Weighted_Results.xlsx',
#       'HLA-B_Ebola_Sudan_GP1_Weighted_Results.xlsx',      
#       'HLA-B_Ebola_Zaire_NP_Weighted_Results.xlsx',
#       'HLA-B_Ebola_Sudan_NP_Weighted_Results.xlsx',
#       'HLA-B_SARS_Wuhan-Hu-1_Weighted_Results.xlsx',      
#       'HLA-B_SARS_DeltaAY4_Weighted_Results.xlsx',
#       'HLA-B_SARS_OmicronBA1_Weighted_Results.xlsx',
#       'HLA-B_SARS_OmicronBA2_Weighted_Results.xlsx',
#       'HLA-B_SARS_OmicronBA5_Weighted_Results.xlsx',
#       'HLA-B_Burkholderia_HCP1_Weighted_Results.xlsx']
# fr =  'HLA-B Sigmaj.xlsx'


fa =  'HLA-C Allele Frequencies.xlsx'               # File for Allele Frequency
fs = ['HLA-C_Ebola_Zaire_GP1_Weighted_Results.xlsx',
      'HLA-C_Ebola_Sudan_GP1_Weighted_Results.xlsx',      
      'HLA-C_Ebola_Zaire_NP_Weighted_Results.xlsx',
      'HLA-C_Ebola_Sudan_NP_Weighted_Results.xlsx',
      'HLA-C_SARS_Wuhan-Hu-1_Weighted_Results.xlsx',      
      'HLA-C_SARS_DeltaAY4_Weighted_Results.xlsx',
      'HLA-C_SARS_OmicronBA1_Weighted_Results.xlsx',
      'HLA-C_SARS_OmicronBA2_Weighted_Results.xlsx',
      'HLA-C_SARS_OmicronBA5_Weighted_Results.xlsx',
      'HLA-C_Burkholderia_HCP1_Weighted_Results.xlsx']
fr =  'HLA-C Sigmaj.xlsx'


lb = ['Ebola GP1 (Zaire)',
      'Ebola GP1 (Sudan)',
      'Ebola NP (Zaire)',
      'Ebola NP (Sudan)',
      'SARS-CoV-2 Wuhan-Hu-1',
      'SARS-CoV-2 Delta AY.4',
      'SARS-CoV-2 Omicron BA.1',
      'SARS-CoV-2 Omicron BA.2',
      'SARS-CoV-2 Omicron BA.5', 
      'Burkholderia HCP1']


# NORMALIZATION OPTIONS:
en = 2 # Enrichment Normalization: 1=Rescale (Old Way) 2=Divide by Sum (New Way)
im = 2 # Immunogenicity: 1=Do Not Use, 2=Use
ft = 2 # Frequency Type: 1=Sum, 2=Average


#%% PARAMETERS ################################################################
    

# Log Enrichment Scores
le = [0.127, #A
     -0.175, #C
      0.072, #D
      0.325, #E
      0.380, #F
      0.110, #G
      0.105, #H
      0.432, #I
     -0.700, #K
     -0.036, #L
     -0.570, #M
     -0.021, #N
     -0.036, #P
     -0.376, #Q
      0.168, #R
     -0.537, #S
      0.126, #T
      0.134, #V
      0.719, #W
     -0.012] #Y
    
    
# Position Importance
pi = [0,    #1
      0,    #2
      0.1,  #3
      0.31, #4
      0.3,  #5
      0.29, #6
      0.26, #7
      0.18, #8
      0]    #9


#%% NORMALIZATION #############################################################


# NORMALIZE POSITION IMPORTANCE:
S = sum(pi)
for i in range(len(pi)):
    pi[i] = pi[i]/S                 # Normalize so that sum(p[i])=1


# NORMALIZE ENRICHMENT SCORES:
for i in range(len(le)):
    le[i] = exp(le[i])              # Undo Natural Log
M = max(le); m = min(le)            # M = Max Enrichment, m = Min Enrichment
S = sum(le)
for i in range(len(le)):        
    if en==1:       
        le[i] = (le[i] - m) / (M-m) # Rescale Enrichment Scores
    if en==2:
        le[i] = le[i]/S             # Normalize Enrichment Scores
        
        
# CREATE ENRICHMENT SCORE DICTIONARY:
aa = list('ACDEFGHIKLMNPQRSTVWY')
es = dict(zip(aa,le))


#%% LOAD ALLELE FREQUENCY DATA ################################################


a  = load_workbook(fd+fa)
rs = a.sheetnames
nr = len(rs)
af = [{} for i in range(nr)]
Z  = []
for i in range(nr):
    ad = []
    if ft==1: m1=18; m2=19
    if ft==2: m1=12; m2=13
    for k in a[rs[i]].iter_rows(min_row=3, min_col=m1, max_col=m2, values_only=True):
        if isinstance(k[0], str): ad.append(k)
        else: break
    na = len(ad)
    S = 0
    for k in range(na):
        af[i]['HLA-'+ad[k][0]]=ad[k][1]
        S = S + ad[k][1]
    Z.append(S)
    for k in af[i]:
        af[i][k] = af[i][k]/S
        
        
#%% MAIN ######################################################################


# DETERMINE NUMBER OF DISEASES:
if isinstance(fs,str): # If fs is a single string
    nd = 1             # Then the number of diseases is 1
else:                  # If fs is a list
    nd = len(fs)       # Then the number of diseases is length of list


up  = []
sj  = [[[] for j in range(nr)] for i in range(nd)]
upo = [[[] for j in range(nr)] for i in range(nd)]
sjo = [[[] for j in range(nr)] for i in range(nd)]


for i in range(nd):
    
    
    # LOAD DATA FOR DISEASE i
    d = load_workbook(fd+fs[i])
    
    
    for j in range(nr):
        
        
        # GET DATA FOR REGION j:
        a1 = []; b1 = []; p1 = [];
        for k in d[rs[j]].iter_rows(min_row=2, values_only=True):
            if isinstance(k[0],str): # If the entry is not blank, Then
                a1.append(k[0]) # Allele Name
                b1.append(k[6]) # Binding Affinity
                p1.append(k[5]) # Peptide Name
        MN = len(a1) # Total Number of Alleles/Peptides
        
        
        # COMPUTE SIGMA j VALUES:
        if j==0:
            up.append(list(unique(p1)));
            N  = len(up[i])
        for k in range(N):
            ix = [l for l in range(MN) if p1[l]==up[i][k]]
            sj[i][j].append(hf.compute_sigma([b1[l] for l in ix],[p1[l] for l in ix],es,pi))
                
    
        sjo[i][j], upo[i][j] = zip(*sorted(zip(sj[i][j],up[i]), reverse=True))
    
 
#%% WRITE RESULTS TO FILE #####################################################


r = load_workbook(fdr+fr)

for i in range(nd):
    
    w = r[lb[i]]
    
    for j in range(nr):
        for k in range(len(sjo[i][j])):
            
            w.cell(k+4,2+2*j).value = upo[i][j][k]
            w.cell(k+4,3+2*j).value = sjo[i][j][k]
            
r.save(fdr+fr)