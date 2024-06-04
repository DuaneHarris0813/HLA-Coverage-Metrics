#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Mar  8 11:33:11 2024

@author: duaneharris
"""


# IMPORT PACKAGES:
from numpy import unique
from openpyxl import load_workbook
import hla_functions as hf


# DATA FILES:
fd  = '../Data/'
fdr = '../Results/'


# OPTIONAL DATA FILES:

# fs = ['HLA-A_Ebola_Zaire_GP1_Weighted_Results.xlsx',
#       'HLA-A_Ebola_Sudan_GP1_Weighted_Results.xlsx',      
#       'HLA-A_Ebola_Zaire_NP_Weighted_Results.xlsx',
#       'HLA-A_Ebola_Sudan_NP_Weighted_Results.xlsx',
#       'HLA-A_SARS_Wuhan-Hu-1_Weighted_Results.xlsx',      
#       'HLA-A_SARS_DeltaAY4_Weighted_Results.xlsx',
#       'HLA-A_SARS_OmicronBA1_Weighted_Results.xlsx',
#       'HLA-A_SARS_OmicronBA2_Weighted_Results.xlsx',
#       'HLA-A_SARS_OmicronBA5_Weighted_Results.xlsx',
#       'HLA-A_Burkholderia_HCP1_Weighted_Results.xlsx']
# fr =  'HLA-A Phij.xlsx'

# fs = ['HLA-B_Ebola_Zaire_GP1_Weighted_Results.xlsx',
#       'HLA-B_Ebola_Sudan_GP1_Weighted_Results.xlsx',      
#       'HLA-B_Ebola_Zaire_NP_Weighted_Results.xlsx',
#       'HLA-B_Ebola_Sudan_NP_Weighted_Results.xlsx',
#       'HLA-B_SARS_Wuhan-Hu-1_Weighted_Results.xlsx',      
#       'HLA-B_SARS_DeltaAY4_Weighted_Results.xlsx',
#       'HLA-B_SARS_OmicronBA1_Weighted_Results.xlsx',
#       'HLA-B_SARS_OmicronBA2_Weighted_Results.xlsx',
#       'HLA-B_SARS_OmicronBA5_Weighted_Results.xlsx',
#       'HLA-B_Burkholderia_HCP1_Weighted_Results.xlsx']
# fr =  'HLA-B Phij.xlsx'

fs = ['HLA-C_Ebola_Zaire_GP1_Weighted_Results.xlsx',
      'HLA-C_Ebola_Sudan_GP1_Weighted_Results.xlsx',      
      'HLA-C_Ebola_Zaire_NP_Weighted_Results.xlsx',
      'HLA-C_Ebola_Sudan_NP_Weighted_Results.xlsx',
      'HLA-C_SARS_Wuhan-Hu-1_Weighted_Results.xlsx',      
      'HLA-C_SARS_DeltaAY4_Weighted_Results.xlsx',
      'HLA-C_SARS_OmicronBA1_Weighted_Results.xlsx',
      'HLA-C_SARS_OmicronBA2_Weighted_Results.xlsx',
      'HLA-C_SARS_OmicronBA5_Weighted_Results.xlsx',
      'HLA-C_Burkholderia_HCP1_Weighted_Results.xlsx']
fr =  'HLA-C Phij.xlsx'


#%% PARAMETERS ################################################################
    

# PATHOGEN LABELS / SHEETNAMES:
lb = ['Ebola GP1 (Zaire)',
      'Ebola GP1 (Sudan)',
      'Ebola NP (Zaire)',
      'Ebola NP (Sudan)',
      'SARS-CoV-2 Wuhan-Hu-1',
      'SARS-CoV-2 Delta AY.4',
      'SARS-CoV-2 Omicron BA.1',
      'SARS-CoV-2 Omicron BA.2',
      'SARS-CoV-2 Omicron BA.5', 
      'Burkholderia HCP1']


# REGIONS:
rs = ['Australia', 
      'Europe',
      'North Africa',
      'North America',
      'North East Asia',
      'Oceania',
      'South and Central America',
      'South Asia',
      'South East Asia',
      'Sub-Saharan Africa',
      'Western Asia']


# GET POSITION WEIGHTS AND ENRICHMENT SCORES:
pw = hf.compute_position_weights()  # Position Weights
es = hf.compute_enrichment_scores() # Enrichment Scores


# SIZES:
nd = len(lb)
nr = len(rs)


#%% MAIN ######################################################################


# INITIALIZE VARIABLES:
up  = []
sj  = [[[] for j in range(nr)] for i in range(nd)]
upo = [[[] for j in range(nr)] for i in range(nd)]
sjo = [[[] for j in range(nr)] for i in range(nd)]


# COMPUTE PHI VALUES:
for i in range(nd):
    
    
    # LOAD DATA FOR DISEASE i
    d = load_workbook(fd+fs[i])
    
    
    for j in range(nr):
        
        
        # GET DATA FOR REGION j:
        a1 = []; b1 = []; p1 = [];
        for k in d[rs[j]].iter_rows(min_row=2, values_only=True):
            if isinstance(k[0],str): # If the entry is not blank, Then
                a1.append(k[0]) # Allele Name
                b1.append(k[6]) # Binding Affinity
                p1.append(k[5]) # Peptide Name
        MN = len(a1) # Total Number of Alleles/Peptides
        
        
        # COMPUTE SIGMA j VALUES:
        if j==0:
            up.append(list(unique(p1)));
            N  = len(up[i])
        for k in range(N):
            ix = [l for l in range(MN) if p1[l]==up[i][k]]
            sj[i][j].append(hf.compute_sigma([b1[l] for l in ix],[p1[l] for l in ix],es,pw))
                
    
        sjo[i][j], upo[i][j] = zip(*sorted(zip(sj[i][j],up[i]), reverse=True))
    
 
#%% WRITE RESULTS TO FILE #####################################################


r = load_workbook(fdr+fr)

for i in range(nd):
    
    w = r[lb[i]]
    
    for j in range(nr):
        for k in range(len(sjo[i][j])):
            
            w.cell(k+4,2+2*j).value = upo[i][j][k]
            w.cell(k+4,3+2*j).value = sjo[i][j][k]
            
r.save(fdr+fr)