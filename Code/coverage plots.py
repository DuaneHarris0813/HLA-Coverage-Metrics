#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Apr  1 11:09:00 2024

@author: duaneharris
"""


# IMPORT LIBRARIES:
from numpy import linspace
from numpy import array
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import seaborn as sb


# DATA FILES:
fdr = '../Results/'
fr1 = 'Coverage Results.xlsx'
fr2 = 'Individual Results'


# ALLELE TYPES:
at = ['HLA-A',
      'HLA-B',
      'HLA-C']


# PATHOGENS:
lb = ['Ebola GP1 (Zaire)',
      'Ebola GP1 (Sudan)',
      'Ebola NP (Zaire)',
      'Ebola NP (Sudan)',
      'SARS-CoV-2 Wuhan-Hu-1',
      'SARS-CoV-2 Delta AY.4',
      'SARS-CoV-2 Omicron BA.1',
      'SARS-CoV-2 Omicron BA.2',
      'SARS-CoV-2 Omicron BA.5', 
      'Burkholderia HCP1']


# REGIONS:
rg = ['Australia', 
      'Europe',
      'North Africa',
      'North America',
      'North East Asia',
      'Oceania',
      'South and Central America',
      'South Asia',
      'South East Asia',
      'Sub-Saharan Africa',
      'Western Asia']


dbu = [0, 4, 9, 10]
dbl = ['Ebola','SARS-CoV-2','Burkholderia']


# VISUAL OPTIONS:
rcs = 1   # Coverage (By Region) Color Scheme
pcs = 2   # Coverage (By Pathogen) Color Scheme
sc  = 0.9 # Brightness Scale


# FIGURE SIZE OPTIONS:
fw  = 504
fh  = 504/3
ppi = 72


# VALUE RANGE OPTIONS:
ta   = 25 # Top Alleles
ym1  = 0.0012
ymf1 = 0.4
ymsg = 0.0025
ym2  = 0.14
ym3  = 0.0025


#%% PARAMETERS ################################################################


# SIZES:
nt = len(at)
nd = len(lb)
nr = len(rg)


# ROW POSTITIONS:
r1 = [4,18,32]
r2 = [4,33,62]


# Color Schemes
cr1 = [[ 1   , 0    , 1   ],
       [ 1   , 0.4  , 1   ],
       [ 1   , 0    , 0   ],
       [ 1   , 0.4  , 0.4 ],
       [ 0   , 1    , 1   ],
       [ 0   , 0.75 , 1   ],
       [ 0   , 0.5  , 1   ],
       [ 0   , 0.25 , 1   ],
       [ 0   , 0    , 1   ],
       [ 0   , 1    , 0   ]]
cr2 = [[0.0, 0.0, 1.0],
       [1.0, 0.5, 0.0],
       [0.0, 1.0, 0.0],
       [1.0, 0.0, 0.0],
       [1.0, 0.5, 0.5],
       [1.0, 0.0, 1.0],
       [1.0, 0.5, 1.0],
       [0.5, 0.5, 0.5],
       [1.0, 1.0, 0.0],
       [0.0, 1.0, 1.0],
       [0.0, 0.5, 1.0]]


#%% READ DATA %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%


# INITIALIZE VARIABLES:
c1r = [[[] for i in range(nd)] for t in range(nt)]
c1p = []
an  = [[[] for j in range(nr)] for t in range(nt)]
f1  = [[[] for j in range(nr)] for t in range(nt)]
sg  = [[[[] for j in range(nr)] for i in range(nd)] for t in range(nt)]
f2  = [[[[] for k in range(ta)] for j in range(nr)] for t in range(nt)]
c2  = [[[[[] for k in range(ta)] for j in range(nr)] for i in range(nd)] for t in range(nt)]


# REGIONAL COVERAGE RESULTS:
r = load_workbook(fdr+fr1)
w = r['Coverage']
for t in range(3):
    for j in w.iter_rows(min_row=r1[t],max_row=r1[t]+nr-1,min_col=3,max_col=12,values_only=True):
        for i in range(nd):
            c1r[t][i].append(j[i])
    c1p.append(list(map(list,zip(*c1r[t])))) # Reorganize Regional Coverage Scores by Pathogen


# FREQ VS SIGMA RESULTS:
for j in range(nr):    
    w = r[rg[j]]
    for t in range(nt):
        for k in w.iter_rows(min_row=r2[t],max_row=r2[t]+ta-1,min_col=2,max_col=13,values_only=True):
            if k[0]==None: an[t][j].append('')
            else: an[t][j].append(k[0])
            f1[t][j].append(k[1])
            for i in range(nd):
                sg[t][i][j].append(k[i+2])
                
                
# INDIVIDUAL FREQUENCY AND COVERAGE:
for j in range(nr):
    r = load_workbook(fdr+fr2+' '+rg[j]+'.xlsx')
    w = r['Frequency']
    for t in range(nt):
        for k,l in enumerate(w.iter_rows(min_row=r2[t],max_row=r2[t]+ta-1,min_col=3,max_col=27,values_only=True)):
            f2[t][j][k] = l
    for i in range(nd):
        w = r[lb[i]]
        for t in range(nt):
            for k,l in enumerate(w.iter_rows(min_row=r2[t],max_row=r2[t]+ta-1,min_col=3,max_col=27,values_only=True)):
                c2[t][i][j][k] = l


#%% SUM INDIVIDUAL FREQUENCIES ################################################
    
    
ac =[[] for t in range(nt)]
for t in range(nt):
    for j in range(nr):
        s=0
        for k in range(ta):
            for l in range(ta):
                if k>=l and f2[t][j][k][l]>0:
                    s = s + f2[t][j][k][l]
        ac[t].append(100*s)
        

#%% CHANGE LABEL NAMES ########################################################
    

# REGION NAMES:
rgl = rg[:]
for j in range(nr):
    rgl[j] = rgl[j].replace('and','&')
    rgl[j] = rgl[j].replace('North East','Northeast')
    rgl[j] = rgl[j].replace('South East','Southeast')
rgs = rgl[:]
for j in range(nr):
    rgs[j] = rgs[j].replace('Northeast','NE')
    rgs[j] = rgs[j].replace('North','N')
    rgs[j] = rgs[j].replace('Southeast','SE')
    rgs[j] = rgs[j].replace('South','S')
    rgs[j] = rgs[j].replace('Western','W')
    rgs[j] = rgs[j].replace('Central','C')
    rgs[j] = rgs[j].replace('Sub-Saharan','SS')
    rgs[j] = rgs[j].replace(' & ','/')
 
    
# PAHTOGEN NAMES: 
lbl = lb[:]
for i in range(nd):
    lbl[i] = lbl[i].replace('GP1','GP')
lbs = lbl[:]
for i in range(nd):
    lbs[i] = lbs[i].replace('Ebola','EBV')
    lbs[i] = lbs[i].replace('Zaire','Z')
    lbs[i] = lbs[i].replace('Sudan','S')
    lbs[i] = lbs[i].replace('SARS-CoV-2','SARS')
    lbs[i] = lbs[i].replace('Wuhan-Hu-1','WH-1')
    lbs[i] = lbs[i].replace('Delta AY.4','D.AY.4')
    lbs[i] = lbs[i].replace('Omicron ','O.')
    lbs[i] = lbs[i].replace('Burkholderia','Burk')
    
 
# ALLELE NAME:
an0 = [[[] for j in range(nr)] for t in range(nt)]
for t in range(nt):
    for j in range(nr):
        for k in range(ta):
            if an[t][j][k]=='':
                an0[t][j].append(an[t][j][k])
            else:
                an0[t][j].append(an[t][j][k][6:])
            
        
#%% REGIONAL COVERAGE BAR PLOT (BY REGION) ####################################


# X-LOCATIONS:
bw = 1/(nd+2)
bx = [[] for i in range(nd)]
bx[0] = [i for i in range(nr)]
for i in range(1,nd):
    bx[i] = [j+bw for j in bx[i-1]]


# SET COLOR SCHEME:
cr = cr2
if rcs == 1: cr = cr1


for t in range(nt):
    
    # PLOT RESULTS:
    fig,ax = plt.subplots(figsize=(fw/ppi,fh/ppi),dpi=300)   
    for i in range(nd):    
        if isinstance(lbs,str):
            plt.bar(bx[i], c1r[t][i], width=bw, label=lbl)
        else:
            plt.bar(bx[i], c1r[t][i], width=bw, label=lbl[i], color=[j * sc for j in cr[i]])
    
    # ADJUST X AXIS:
    if t==2:
        plt.xlabel('Regions', fontweight='bold')
    plt.xticks([i+bw*(nd-1)/2 for i in range(nr)], rgs, rotation=45, ha='right', fontsize=8)    
    if t!=2:
        ax.tick_params(labelbottom=False)   
      
    # ADJUST Y AXIS    
    plt.ylabel(r'$C_k$', fontweight='bold')   
    plt.ylim((0,ym1))
    plt.ticklabel_format(axis="y", style="sci", scilimits=(0,0))
    ax.yaxis.major.formatter._useMathText = True
    
    # CREATE LEGEND:
    if t==2:
        plt.legend(fontsize=8, bbox_to_anchor=(0,-0.5), loc='upper left', ncol=3)
    
    # SAVE PLOTS:
    plt.savefig(fdr+'Figures/Vaccine Protein Coverage Metrics/Ck by Region/'+at[t]+' Ck by Region.pdf',format='pdf',bbox_inches='tight')
    plt.show()


#%% REGIONAL COVERAGE BAR PLOT (BY PATHOGEN) ##################################


# X-LOCATIONS:
bw = 1/(nr+2)
bx = [[] for i in range(nr)]
bx[0] = [i for i in range(nd)]
for i in range(1,nr):
    bx[i] = [j+bw for j in bx[i-1]]
    
    
# SET COLOR SCHEME:
cr = cr2
if pcs == 1: cr = cr1
    
    
for t in range(nt):
    
    # PLOT RESULTS:
    fig,ax = plt.subplots(figsize=(fw/ppi,fh/ppi),dpi=300)   
    for j in range(nr):    
        if isinstance(lbs,str):
            plt.bar(bx[j], c1p[t][j], width=bw, label=rgl)
        else:
            plt.bar(bx[j], c1p[t][j], width=bw, label=rgl[j], color=[j * sc for j in cr[j]])
        
    # ADJUST X AXIS:
    if t==2:
        plt.xlabel('Pathogens', fontweight='bold')
    plt.xticks([i+bw*(nr-1)/2 for i in range(nd)], lbs, rotation=45, ha='right', fontsize=8)    
    if t!=2:
        ax.tick_params(labelbottom=False)    
        
    # ADJUST Y AXIS    
    plt.ylabel(r'$C_k$', fontweight='bold')
    plt.ylim((0,ym1))
    plt.ticklabel_format(axis="y", style="sci", scilimits=(0,0))
    ax.yaxis.major.formatter._useMathText = True
    
    # CREATE LEGEND:
    if t==2:
        plt.legend(fontsize=8, bbox_to_anchor=(0,-0.5), loc='upper left', ncol=3)
    
    # SAVE PLOTS:
    plt.savefig(fdr+'Figures/Vaccine Protein Coverage Metrics/Ck by Pathogen/'+at[t]+' Ck by Pathogen.pdf',format='pdf',bbox_inches='tight')
    plt.show()
    
    
#%% FREQ VS SIGMA PLOTS #######################################################


for i in range(len(dbu)-1):
    
    ndi = dbu[i+1] - dbu[i]
    bw = 1/(ndi+3)
    
    bx = [[] for k in range(ndi+1)]
    bx[0] = [k for k in range(ta)]
    for k in range(1,ndi+1):
        bx[k] = [l+bw for l in bx[k-1]]
    
    for j in range(nr): 
        for t in range(nt):
        
            # PLOT RESULTS:
            fig,ax1 = plt.subplots(figsize=(fw/ppi,fh/ppi),dpi=300)   
            ax2 = ax1.twinx()
            ax1.bar(bx[0], f1[t][j], width=bw, color='black', label="Frequency")
            for k in range(1,ndi+1):
                if isinstance(lbs,str):
                    ax2.bar(bx[k], sg[t][k-1][j], width=bw)
                else:
                    ax2.bar(bx[k], sg[t][k-1][j], width=bw, color=cr1[k-1+dbu[i]], label=lbs[dbu[i]+k-1])
            
            # ADJUST X AXIS:
            if t==2:
                ax1.set_xlabel('Alleles',fontweight='bold')
            ax1.set_xticks([k+bw*ndi/2 for k in range(ta)], an0[t][j], rotation=90, fontsize=8)    
            
            # ADJUST Y AXIS:
            ax1.set_ylabel(r'$\hat{f}$', fontweight='bold')          
            ax1.set_ylim((0,ymf1))
            ax2.set_ylabel(r'$\sigma$', fontweight='bold')
            ax2.set_ylim((0,ymsg))
            plt.ticklabel_format(axis="y", style="sci", scilimits=(0,0))
            ax2.yaxis.major.formatter._useMathText = True
        
            # CREATE LEGENDS:
            if t==2:
                ax1.legend(fontsize=8, loc='upper left')
                ax2.legend(fontsize=8, loc='upper right', ncol=2)
            
            # SAVE PLOTS:
            plt.savefig(fdr+'Figures/Frequency vs Sigma/'+dbl[i]+'/'+rgl[j]+'/'+at[t]+' FvS '+dbl[i]+' '+rg[j]+'.pdf',format='pdf',bbox_inches='tight')
            plt.show()
    

#%% INDIVIDUAL FREQUENCY HEAT MAPS ############################################
    

for j in range(nr):
    for t in range(nt):

        # PLOT RESULTS:      
        fig,ax = plt.subplots(figsize=(fw/ppi,0.65*fw/ppi),dpi=300)
        data = array(f2[t][j])
        sb.heatmap(data,mask=data<0,vmin=0,vmax=ym2,cmap='hot',cbar=False)
        
        # SET AXIS LABELS:
        ax.set_xticks(linspace(0,ta-1,ta)+0.5,an0[t][j],rotation=90,fontsize=12)
        ax.set_yticks(linspace(0,ta-1,ta)+0.5,an0[t][j],rotation=0,fontsize=12)
        
        # ANNOTATE POPULATION COVERAGE:
        plt.annotate('%.0f%% of the Population' % ac[t][j], xy=(12,3), fontsize=12)
        
        # SAVE PLOTS:
        plt.savefig(fdr+'Figures/Individual Coverage Heat Maps/Frequencies/'+rgl[j]+'/'+at[t]+' Individual Frequencies '+rgl[j]+'.pdf',format='pdf',bbox_inches='tight')
        plt.show()
        
        # CREATE COLOR BAR:
        if j==0 and t==0:
            fig,ax = plt.subplots(figsize=(fw/ppi,fw/ppi),dpi=300)
            sb.heatmap(data,mask=data<0,vmin=0,vmax=ym2,cmap='hot',cbar_kws={'orientation': 'horizontal', 'label': 'Allele Frequency'})
            ax.remove()
            plt.savefig(fdr+'Figures/Individual Coverage Heat Maps/Individual Frequency Color Map.pdf',format='pdf',bbox_inches='tight')


#%% INDIVIDUAL COVERAGE HEAT MAPS #############################################


for j in range(nr):
    for i in range(nd):   
        for t in range(nt):
      
            # PLOT RESULTS:
            fig,ax = plt.subplots(figsize=(fw/ppi,0.65*fw/ppi),dpi=300)
            data = array(c2[t][i][j])
            sb.heatmap(data,mask=data<=0,vmin=0,vmax=ym3,cmap='hot',cbar=False)
            
            # SET AXIS LABELS
            ax.set_xticks(linspace(0,ta-1,ta)+0.5,an0[t][j],rotation=90,fontsize=12)
            ax.set_yticks(linspace(0,ta-1,ta)+0.5,an0[t][j],rotation=0,fontsize=12)

            # SAVE PLOTS:
            plt.savefig(fdr+'Figures/Individual Coverage Heat Maps/'+lb[i]+'/'+rgl[j]+'/'+at[t]+' Individual Coverage '+lb[i]+' '+rgl[j]+'.pdf',format='pdf',bbox_inches='tight')
            plt.show()
            
            # CREATE COLOR BAR:
            if i==0 and j==0 and t==0:
                fig,ax = plt.subplots(figsize=(fw/ppi,0.65*fw/ppi),dpi=300)
                sb.heatmap(data,mask=data<=0,vmin=0,vmax=ym3,cmap='hot',cbar_kws={'orientation': 'horizontal', 'label': 'Individual Coverage Score'})
                ax.remove()
                plt.savefig(fdr+'Figures/Individual Coverage Heat Maps/Individual Coverage Color Map.pdf',format='pdf',bbox_inches='tight')
                plt.show()