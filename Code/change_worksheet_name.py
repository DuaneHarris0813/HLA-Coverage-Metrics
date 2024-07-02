#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jun 28 13:41:00 2024

@author: duaneharris
"""

from os import walk
from openpyxl import load_workbook

fd = '../Results/'

pno = 'Burkholderia Hcp11'
pnn = 'Burkholderia Hcp1'

fs = []
for r,_,n in walk(fd):
    for f in n:
        if f.lower().endswith('.xlsx'):
            fs.append(f)
            
for f in fs:
    if f[0]!='~':
        d = load_workbook(fd+f)
        s = d.sheetnames
        if pno in s:
            w = d[pno]
            w.title = pnn
            d.save(fd+f)