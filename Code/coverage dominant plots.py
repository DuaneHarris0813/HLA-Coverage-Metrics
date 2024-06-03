#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Apr 18 14:47:33 2024

@author: duaneharris
"""


# IMPORT LIBRARIES:
from openpyxl import load_workbook
import matplotlib.pyplot as plt


# DATA FILES:
fdr = '../Results/'           # Results Folder
fr1 = 'Coverage Results.xlsx' # Results File


# ALLELE TYPES:
at = ['HLA-A',
      'HLA-B',
      'HLA-C']


# PATHOGENS:
lb = ['Ebola GP1 (Zaire)',
      'Ebola GP1 (Sudan)',
      'SARS-CoV-2 Wuhan-Hu-1',
      'SARS-CoV-2 Delta AY.4',
      'SARS-CoV-2 Omicron BA.1',
      'SARS-CoV-2 Omicron BA.2',
      'SARS-CoV-2 Omicron BA.5']


# REGIONS:
rg = ['Australia', 
      'Europe',
      'North Africa',
      'North America',
      'North East Asia',
      'Oceania',
      'South and Central America',
      'South Asia',
      'South East Asia',
      'Sub-Saharan Africa',
      'Western Asia']


# VISUAL OPTIONS:
rcs = 1   # Coverage (By Region) Color Scheme
pcs = 2   # Coverage (By Pathogen) Color Scheme
sc  = 0.9 # Brightness Scale


# FIGURE SIZE OPTIONS:
fw  = 504   # Figure Width
fh  = 504/3 # Figure Height
ppi = 72    # Point Per Inch (Conversion Ratio)


# VALUE RANGE OPTIONS:
ym = 0.09 # Upper Y Limit


#%% PARAMETERS ################################################################


# SIZES:
nt = len(at) # Number of Allele Types
nd = len(lb) # Number of Pathogens
nr = len(rg) # Number of Regions


# ROW POSTITIONS:
r1 = [4,18,32]


# COLOR SCHEMES:
cr1 = [[1.00, 0.00, 1.00],
       [1.00, 0.40, 1.00],
       [0.00, 1.00, 1.00],
       [0.00, 0.75, 1.00],
       [0.00, 0.50, 1.00],
       [0.00, 0.25, 1.00],
       [0.00, 0.00, 1.00],
       [0.00, 1.00, 0.00]]
cr2 = [[0.00, 0.00, 1.00],
       [1.00, 0.50, 0.00],
       [0.00, 1.00, 0.00],
       [1.00, 0.00, 0.00],
       [1.00, 0.50, 0.50],
       [1.00, 0.00, 1.00],
       [1.00, 0.50, 1.00],
       [0.50, 0.50, 0.50],
       [1.00, 1.00, 0.00],
       [0.00, 1.00, 1.00],
       [0.00, 0.50, 1.00]]


#%% READ DATA %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%


# INITIALIZE VARIABLES:
Fr  = [[[] for i in range(nd)] for t in range(nt)]
Fp  = []

    
# REGIONAL COVERAGE RATIOS:
r = load_workbook(fdr+fr1)
w = r['Dominant Coverage']
for t in range(3):
    for j in w.iter_rows(min_row=r1[t],max_row=r1[t]+nr-1,min_col=3,max_col=9,values_only=True):
        for i in range(nd):
            Fr[t][i].append(j[i])
    Fp.append(list(map(list,zip(*Fr[t]))))
        

#%% CHANGE LABEL NAMES ########################################################
    

# REGION NAMES:
rgl = rg[:]
for j in range(nr):
    rgl[j] = rgl[j].replace('and','&')
    rgl[j] = rgl[j].replace('North East','Northeast')
    rgl[j] = rgl[j].replace('South East','Southeast')
rgs = rgl[:]
for j in range(nr):
    rgs[j] = rgs[j].replace('Northeast','NE')
    rgs[j] = rgs[j].replace('North','N')
    rgs[j] = rgs[j].replace('Southeast','SE')
    rgs[j] = rgs[j].replace('South','S')
    rgs[j] = rgs[j].replace('Western','W')
    rgs[j] = rgs[j].replace('Central','C')
    rgs[j] = rgs[j].replace('Sub-Saharan','SS')
    rgs[j] = rgs[j].replace(' & ','/')
 
    
# PAHTOGEN NAMES: 
lbl = lb[:]
for i in range(nd):
    lbl[i] = lbl[i].replace('GP1','GP')
    lbl[i] = lbl[i].replace('HCP1','Hcp1')
lbs = lbl[:]
for i in range(nd):
    lbs[i] = lbs[i].replace('Ebola','EBV')
    lbs[i] = lbs[i].replace('Zaire','Z')
    lbs[i] = lbs[i].replace('Sudan','S')
    lbs[i] = lbs[i].replace('SARS-CoV-2','SARS')
    lbs[i] = lbs[i].replace('Wuhan-Hu-1','WH-1')
    lbs[i] = lbs[i].replace('Delta AY.4','D.AY.4')
    lbs[i] = lbs[i].replace('Omicron ','O.')
    lbs[i] = lbs[i].replace('Burkholderia','Burk')
    
    
#%% COVERAGE RATIOS (BY REGION) ###############################################


# X-LOCATIONS:
bw = 1/(nd+2)
bx = [[] for i in range(nd)]
bx[0] = [i for i in range(nr)]
for i in range(1,nd):
    bx[i] = [j+bw for j in bx[i-1]]
    
    
# SET COLOR SCHEME:
cr = cr2
if rcs == 1: cr = cr1


for t in range(nt):
    
    # PLOT RESULTS:
    fig,ax = plt.subplots(figsize=(fw/ppi,fh/ppi),dpi=300)   
    for i in range(nd):    
        if isinstance(lbs,str):
            plt.bar(bx[i], Fr[t][i], width=bw, label=lbl)
        else:
            plt.bar(bx[i], Fr[t][i], width=bw, label=lbl[i], color=[j * sc for j in cr[i]])
    
    # ADJUST X AXIS:
    if t==2:
        plt.xlabel('Regions', fontweight='bold')
    plt.xticks([i+bw*(nd-1)/2 for i in range(nr)], rgs, rotation=45, ha='right', fontsize=8)    
    if t!=2:
        ax.tick_params(labelbottom=False) 
        
    # ADJUST Y AXIS    
    plt.ylabel(r'$F_k$', fontweight='bold')
    plt.ylim((0,ym))
    plt.ticklabel_format(axis="y", style="sci", scilimits=(0,0))
    ax.yaxis.major.formatter._useMathText = True    
    
    # CREATE TITLE:
    plt.title(at[t])
    
    # CREATE LEGEND:
    if t==2:
        plt.legend(fontsize=8, bbox_to_anchor=(0,-0.5), loc='upper left', ncol=2)
       
    # SAVE PLOTS:
    plt.savefig(fdr+'Figures/Coverage Ratios/Fk by Region/'+at[t]+' Fk by Region.pdf',format='pdf',bbox_inches='tight')
    plt.show()   
    

#%% COVERAGE RATIOS (BY PATHOGEN) #############################################
    

# X-LOCATIONS:
bw = 1/(nr+2)
bx = [[] for i in range(nr)]
bx[0] = [i for i in range(nd)]
for i in range(1,nr):
    bx[i] = [j+bw for j in bx[i-1]]
    
    
# SET COLOR SCHEME:
cr = cr2
if pcs == 1: cr = cr1
    
    
for t in range(nt):
    
    # PLOT RESULTS:
    fig,ax = plt.subplots(figsize=(fw/ppi,fh/ppi),dpi=300)   
    for j in range(nr):    
        if isinstance(lbs,str):
            plt.bar(bx[j], Fp[t][j], width=bw, label=rgl)
        else:
            plt.bar(bx[j], Fp[t][j], width=bw, label=rgl[j], color=[j * sc for j in cr[j]])
        
    # ADJUST X AXIS:
    if t==2:
        plt.xlabel('Pathogens', fontweight='bold')
    plt.xticks([i+bw*(nr-1)/2 for i in range(nd)], lbs, rotation=45, ha='right', fontsize=8)    
    if t!=2:
        ax.tick_params(labelbottom=False)    
        
    # ADJUST Y AXIS    
    plt.ylabel(r'$F_k$', fontweight='bold')
    plt.ylim((0,ym))
    plt.ticklabel_format(axis="y", style="sci", scilimits=(0,0))
    ax.yaxis.major.formatter._useMathText = True
    
    # CREATE TITLE:
    plt.title(at[t])
    
    # CREATE LEGEND:
    if t==2:
        plt.legend(fontsize=8, bbox_to_anchor=(0,-0.5), loc='upper left', ncol=3)
    
    # SAVE PLOTS:
    plt.savefig(fdr+'Figures/Coverage Ratios/Fk by Pathogen/'+at[t]+' Fk by Pathogen.pdf',format='pdf',bbox_inches='tight')
    plt.show()